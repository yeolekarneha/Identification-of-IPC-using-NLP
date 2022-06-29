



import sys 
import pandas as pd
import numpy as np
import ast
from sklearn.preprocessing import MultiLabelBinarizer
import tensorflow as tf
from tensorflow.keras.layers import TextVectorization
from tensorflow.keras import layers
from keras.models import load_model
from openpyxl import load_workbook
import tensorflow_hub as hub

def IPC_Self_learning_Model(text):
  tf_hub_embedding_layer = hub.KerasLayer("universal-sentence-encoder_4",
                                        trainable=False,
                                        name="universal_sentence_encoder")
  train_data=pd.read_excel("ipc_data 120-309.xlsx")
  X = train_data["Offence"].to_list()
  X = np.array(X)
  train_data_1= train_data['Section']
  train_data_1 = train_data_1.apply(lambda x: ast.literal_eval(x))

  multilabel = MultiLabelBinarizer()
  y = multilabel.fit_transform(train_data_1)

  # Define feature extractor model using TF Hub layer
  inputs = layers.Input(shape=[], dtype=tf.string)
  pretrained_embedding = tf_hub_embedding_layer(inputs) # tokenize text and create embedding
  x = layers.Dense(128, activation="relu")(pretrained_embedding) # add a fully connected layer on top of the embedding
  # Note: you could add more layers here if you wanted to
  outputs = layers.Dense(len(multilabel.classes_), activation="softmax")(x) # create the output layer
  model_3 = tf.keras.Model(inputs=inputs,
                          outputs=outputs)

  # Compile the model
  model_3.compile(loss="categorical_crossentropy",
                  optimizer=tf.keras.optimizers.Adam(),
                  metrics=["accuracy"])

  model_3.fit(X, # input sentences can be a list of strings due to text preprocessing layer built-in model
                                y,
                                epochs=30,
                                verbose=0
                                )
  text = np.array(text)
  model_3_pred_probs = model_3.predict(text)
  n=2
  indices_2 = (-model_3_pred_probs).argsort()[:n]
  q=[]
  for r in range(0,4):
    q.append(multilabel.classes_[indices_2[0,r]])
  

  return q

def IPC_Check(pred,res):
  g=[]
  for r in res:
    if r in pred:
      g.append(True)
    else:
      g.append(False)
  if False in g:
    u="wrong"
  else:
    u="right"
  return u

def Improve(result,off,sec):
  if result=="wrong":

    sec=str(sec)


    myFileName=r'ipc_data 120-309.xlsx'
    #load the workbook, and put the sheet into a variable
    wb = load_workbook(filename=myFileName)
    ws = wb['Sheet1']

    #max_row is a sheet function that gets the last row in a sheet.
    newRowLocation = ws.max_row +1

    #write to the cell you want, specifying row and column, and value :-)
    ws.cell(column=2,row=newRowLocation, value=sec)
    ws.cell(column=3,row=newRowLocation, value=off)
    wb.save(filename=myFileName)
    wb.close()

##Exapme


